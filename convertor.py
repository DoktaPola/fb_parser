# import pandas as pd
#
# read_file = pd.read_excel(r'C:\fb_parser\FB_data.xlsx')
# read_file.to_csv(r'C:\fb_parser\fb_data.csv', index=None, header=True)

import pandas as pd

import openpyxl


def create_csv():
    wb = openpyxl.load_workbook(r'C:\fb_parser\FB_data.xlsx')
    all_dicts = list()
    id = 1

    for ws_name in wb.sheetnames:
        rows_size = wb[ws_name].max_row
        for i in range(2, rows_size):
            link = wb[ws_name].cell(row=i, column=1).hyperlink.target
            name = wb[ws_name].cell(row=i, column=1).value

            d = {'id': id,
                 'name': name,
                 'link': link}

            all_dicts.append(d)
            id += 1

    df = pd.DataFrame(all_dicts)
    df.to_csv(r'fb_data.csv', index=False, header=True)
    wb.close()


create_csv()
